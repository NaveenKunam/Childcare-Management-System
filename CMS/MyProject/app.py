from datetime import datetime
import io
from flask import Flask, Response, render_template, request, redirect, url_for, jsonify
import mysql.connector
import webbrowser

import openpyxl

app = Flask(__name__)
userrole = '' 

db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'Test123',
    'database': 'Childcare_DB',
}

#HomePage starts
@app.route('/', methods=['GET'])
def homepage():
    return render_template('index.html')
#Homepage ends   

#Login Authentication starts
@app.route('/authenticate', methods=['POST'])
def authenticate():
    username = request.form['username']
    password = request.form['password']

    global userrole

    try:
        # Connect to the MySQL database
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Execute an SQL query to retrieve data
        query = f"SELECT * FROM UserLogin where Username= '{username}' and Password= '{password}'"
        cursor.execute(query)

        # Fetch all rows as a list of dictionaries
        data = cursor.fetchall()
        
        userrole = data[0]['Role']

        # Close the cursor and connection
        cursor.close()
        conn.close()

        #return jsonify(data)
        if data:
            return render_template('dashboard.html', user_role = userrole)
        else:
            return render_template('index.html')

    except Exception as e:
        return jsonify({'error': str(e)})
#Login Authentication ends

#SignUp starts
@app.route('/signup', methods =['POST'])
def signup():
    msg = ''
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form and 'email' in request.form and 'role' in request.form:
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        role = request.form['role']
        # Connect to the MySQL database
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Execute an SQL query to retrieve data
        query = f"SELECT * FROM UserLogin where Username= '{username}'"
        cursor.execute(query)
        
        # Fetch all rows as a list of dictionaries
        data = cursor.fetchone()

        if data:
            return jsonify({'message': 'Account already exists !'})
        else:
            # Execute an SQL query to retrieve data
            query1 = "INSERT INTO UserLogin (Username, Password, Email, Role) VALUES (%s, %s, %s, %s)"
            values = (username, password, email, role)

            # Execute the INSERT statement
            cursor.execute(query1, values)

            # Commit the changes to the databases
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

            return render_template('index.html')
#SignUp ends

#Child Enrollment starts -> add_child
@app.route('/add_child', methods=['GET', 'POST'])
def add_child():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get the data from the form
            first_name = request.form['first_name']
            last_name = request.form['last_name']
            dob = request.form['dob']
            allergies = request.form['allergies']
            parents_name = request.form['parents_name']
            address = request.form['address']
            email = request.form['email']
            phone_number = request.form['phone_number']
            consent = request.form.get('consentCheckbox')
            classroom = request.form['classroom']
            fee = calculate_fee(classroom)

            # Construct an SQL INSERT statement
            query = "INSERT INTO children (first_name, last_name, parents_name, email, phone_number, classroom, fee, DOB, Allergies, Address, ConsentFlag) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            values = (first_name, last_name, parents_name, email, phone_number, classroom, fee, dob, allergies, address, consent)

            # Execute the INSERT statement
            cursor.execute(query, values)

            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

        except Exception as e:
            return jsonify({'error': str(e)})

    return render_template('add_child.html', user_role = userrole, success_message="Data loaded successfully!")

def calculate_fee(classroom):
    fees = {
        'Infant': 300,
        'Toddler': 275,
        'Twadler': 250,
        '3 Years Old': 225,
        '4 Years Old': 200
    }
    return fees.get(classroom, 0)
#Child Enrollment ends

@app.route('/teacher_enrollment', methods=['GET', 'POST'])
def teacher_enrollment():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get form data
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')
            dob = request.form.get('DOB')
            Phone = request.form.get('Phone')
            address = request.form.get('address')
            daily_salary = 2000
            salaryperweek = 5000

            # Save data to the database using placeholders
            query = "INSERT INTO teacher (first_name, last_name, dob, phone, address, daily_salary, salary_per_week) VALUES (%s, %s, %s, %s, %s, %s, %s)"
            values = (first_name, last_name, dob, Phone, address, daily_salary,salaryperweek)

            cursor.execute(query, values)
            conn.commit()
            cursor.close()
            conn.close()
    
        except Exception as e:
            return jsonify({'error': str(e)})

    return render_template('add_teacher.html', user_role = userrole, success_message="Data loaded successfully!")

#View_child payment starts
@app.route('/view_child', methods=['GET', 'POST'])
def view_child():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get the data from the form
            child_name = request.form['child_name']

            # Construct an SQL INSERT statement
            query = f"SELECT * FROM children WHERE first_name = '{child_name}'"
            print(query)
            # Execute the INSERT statement
            cursor.execute(query)
            data = cursor.fetchone()
            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

            if data:
                return render_template('view_child.html', child=data, user_role = userrole)
            else:
                return render_template('child_not_found.html', child_name=child_name, user_role = userrole)
    
        except Exception as e:
            return jsonify({'error': str(e)})
#View_child payment ends   

#View_Teacher payment starts
@app.route('/view_teacher', methods=['GET', 'POST'])
def view_teacher():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get the data from the form
            teacher_name = request.form['teacher_name']

            # Construct an SQL INSERT statement
            teacher = "SELECT * FROM teacher WHERE first_name = %s"
            value = (teacher_name)

            # Execute the INSERT statement
            cursor.execute(teacher, value)

            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

            if teacher:
                return render_template('teacher_view_attendance.html', teacher=teacher, user_role = userrole)
            else:
                return render_template('child_not_found.html', teacher_name=teacher_name, user_role = userrole)
    
        except Exception as e:
            return jsonify({'error': str(e)})
#View_child payment ends      

#Fee_Pay starts
@app.route('/pay_fee', methods=['POST'])
def pay_fee():
    if request.method == 'POST':
        child_id = request.form['child_id']
        print(child_id)
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        query = "UPDATE children SET fee = 0 WHERE id = %s"
        values = (child_id,)
        # Execute the INSERT statement
        cursor.execute(query, values)

        # Commit the changes to the database
        conn.commit()

        # Close the cursor and connection
        cursor.close()
        conn.close()
        return jsonify({'message': 'Paid successfully'})

def increase_fees_on_monday():
    if datetime.today().weekday() == 0:
        with app.app_context():
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            fee_increase = {
                'Infant': 300,
                'Toddler': 275,
                'Twadler': 250,
                '3 Years Old': 225,
                '4 Years Old': 200
            }

            for classroom, increase_amount in fee_increase.items():
                query = "UPDATE children SET fee = fee + %s WHERE classroom = %s"
                values = (increase_amount, classroom)
            cursor.execute(query, values)

            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

increase_fees_on_monday()
#Fee_Pay ends

#class registartion start
@app.route('/class_reg')
def class_reg():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM children")
    children_data = cursor.fetchall()

    classroom_details = {}
    for child in children_data:
        classroom = child[6]  
        if classroom not in classroom_details:
            classroom_details[classroom] = {'children_count': 0, 'teacher_count': 0}

        classroom_details[classroom]['children_count'] += 1

        if classroom == 'infant' and classroom_details[classroom]['children_count'] > 4:
            classroom_details[classroom]['teacher_count'] = 2
        elif classroom == 'toddler' and classroom_details[classroom]['children_count'] > 6:
            classroom_details[classroom]['teacher_count'] = 2
        elif classroom == 'twadller' and classroom_details[classroom]['children_count'] > 8:
            classroom_details[classroom]['teacher_count'] = 2
        elif classroom == '3 years' and classroom_details[classroom]['children_count'] > 9:
            classroom_details[classroom]['teacher_count'] = 2
        elif classroom == '4 years' and classroom_details[classroom]['children_count'] > 10:
            classroom_details[classroom]['teacher_count'] = 2

    cursor.execute("SELECT * FROM teacher")
    teachers_data = cursor.fetchall()

    assigned_children = {}
    for teacher in teachers_data:
        teacher_id = teacher[0]
        cursor.execute(f"SELECT * FROM children WHERE teacher_id = {teacher_id}")
        assigned_children[teacher_id] = cursor.fetchall()

    return render_template('class_registration.html', user_role=userrole, classroom_details=classroom_details, teachers_data=teachers_data, assigned_children=assigned_children)
#class reg block end

#Dashboard List starts
@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html', user_role = userrole)
#Dashboard List ends

@app.route('/logout')
def logout():
    return render_template('index.html')

#Facilities enrollment starts
@app.route('/submit_enrollment', methods=['GET', 'POST'])
def insert_data():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get the data from the form
            facilityname = request.form['facilityname']
            address = request.form['address']
            phno = request.form['phno']
            licenseno = request.form['licenseno']
            adminname = request.form['adminname']
            adminemailid = request.form['adminemailid']
            adminphno = request.form['adminphno']
            area = request.form['area']

            # Construct an SQL INSERT statement
            query = "INSERT INTO Facilities_Enrollment (FacilityName, Address, PhoneNumber, LicenseNumber, AdminName, AdminEmail, AdminPhoneNumber, Area ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
            values = (facilityname, address, phno, licenseno, adminname, adminemailid, adminphno, area)

            # Execute the INSERT statement
            cursor.execute(query, values)

            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

            return jsonify({'message': 'Data inserted successfully'})

        except Exception as e:
            return jsonify({'error': str(e)})

    return render_template('attendance.html', user_role = userrole)
#Facilities enrollment ends

#Child_Attendance Entry starts
@app.route('/submit_attendance', methods=['POST'])
def submit_attendance():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get the data from the form
            studentname = request.form['student_name']
            attendancestatus = request.form['attendance_status']
            dateofAttendance = datetime.now()
            #formatted_date = dateofAttendance.strftime("%Y-%m-%d %H:%M:%S")
            # Construct an SQL INSERT statement
            query = "INSERT INTO StudentAttendance (StudentName, Status, Date ) VALUES (%s, %s, %s)"
            values = (studentname, attendancestatus, dateofAttendance)

            # Execute the INSERT statement
            cursor.execute(query, values)
            print(query)
            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

            return jsonify({'message': 'Student Attendance added successfully'})

        except Exception as e:
            return jsonify({'error': str(e)})
#Child_Attendance Entry ends

#View Child_Attendance Summary starts
@app.route('/view_attendance', methods=['GET'])
def view_attendance():
    try:
        # Connect to the MySQL database
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Execute an SQL query to retrieve data
        query = "SELECT * FROM StudentAttendance"
        cursor.execute(query)

        # Fetch all rows as a list of dictionaries
        data = cursor.fetchall()

        # Close the cursor and connection
        cursor.close()
        conn.close()

    except Exception as e:
        return jsonify({'error': str(e)})
    return render_template('view_attendance.html', attendance_data=data, user_role = userrole)
#View Child_Attendance Summary ends

#Teacher_Attendance Entry starts
@app.route('/submit_teacher_attendance', methods=['POST'])
def submit_teacher_attendance():
    if request.method == 'POST':
        try:
            # Connect to the MySQL database
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            # Get the data from the form
            studentname = request.form['teacher_name']
            attendancestatus = request.form['attendance_status']
            dateofAttendance = datetime.now()
            #formatted_date = dateofAttendance.strftime("%Y-%m-%d %H:%M:%S")
            # Construct an SQL INSERT statement
            query = "INSERT INTO TeacherAttendance (TeacherName, Status, Date ) VALUES (%s, %s, %s)"
            values = (studentname, attendancestatus, dateofAttendance)

            # Execute the INSERT statement
            cursor.execute(query, values)
            print(query)
            # Commit the changes to the database
            conn.commit()

            # Close the cursor and connection
            cursor.close()
            conn.close()

            return jsonify({'message': 'Student Attendance added successfully'})

        except Exception as e:
            return jsonify({'error': str(e)})
#Teacher_Attendance Entry ends

#View Teacher_Attendance Summary starts
@app.route('/view_teacher_attendance', methods=['GET'])
def view_teacher_attendance():
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Execute an SQL query to retrieve data
        query = "SELECT * FROM TeacherAttendance"
        cursor.execute(query)

        # Fetch all rows as a list of dictionaries
        data = cursor.fetchall()

        # Close the cursor and connection
        cursor.close()
        conn.close()
        
    except Exception as e:
        return jsonify({'error': str(e)})
    
    return render_template("teacher_view_attendance.html", attendance_data = data, user_role = userrole)

#View Teacher_Attendance Summary ends

#class reg
@app.route('/assign_staff', methods=['POST'])
def assign_staff():
    classroom = request.form.get('classroom')
    staff_name = request.form.get('staff_name')

    return redirect(url_for('db'))

#page routing
@app.route('/page/<page_name>')
def page(page_name):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    if page_name == 'page1':
        # Fetch teacher1 details
        cursor.execute("SELECT * FROM teacher")
        teacher1_data = cursor.fetchall()

        return render_template('page1.html', user_role=userrole, teacher1_data=teacher1_data)
    else:
        return render_template(f'{page_name}.html', user_role=userrole)
#end

#Report Module starts
@app.route('/studentattendance', methods=['GET', 'POST'])
def studentattendance():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    # Initialize variables
    selected_date = None
    classroom = None
    no_records_found = False
    data = []  # Initialize data as an empty list

    if request.method == 'POST':
        # If the form is submitted, get the date and classroom from the form
        selected_date = request.form['date']
        classroom = request.form['classroom']

        # Fetch children attendance data for the selected date and classroom from the 'studentattendance' table
        query = "SELECT sa.StudentName, sa.Status, sa.Date, c.classroom FROM StudentAttendance sa JOIN children c ON sa.StudentID = c.id WHERE sa.Date = %s AND c.classroom = %s"
        cursor.execute(query, (selected_date, classroom))
        data = cursor.fetchall()
        print(data)

        # Variable to track if no records were found
        no_records_found = not data  # Set to True if no records were found

    # Render the template with the attendance data and no_records_found variable
    return render_template('report_student_attendance.html', user_role=userrole, data=data, selected_date=selected_date, classroom=classroom, no_records_found=no_records_found)

@app.route('/download_studentattendance')
def download_studentattendance():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    # Get the date parameter from the URL, defaulting to today's date if not provided
    date_param = request.args.get('date', default=datetime.now().strftime('%Y-%m-%d'))
    classroom = request.args.get('classroom')

    # Query data for the selected date and classroom from the 'studentattendance' table
    query = "SELECT sa.StudentName, sa.Status, sa.Date, c.classroom FROM StudentAttendance sa JOIN children c ON sa.StudentID = c.id WHERE sa.Date = %s AND c.classroom = %s"
    cursor.execute(query, (date_param, classroom))
    data = cursor.fetchall()
    print(data)


    # Variable to track if no records were found
    no_records_found = not data  # Set to True if no records were found

    # Generate Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write header
    for col_num, col_description in enumerate(cursor.description, 1):
        sheet.cell(row=1, column=col_num, value=col_description[0])

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Save the workbook to an in-memory buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Send the Excel file as a response to the client
    return Response(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=output_{date_param}_{classroom}.xlsx"
        }
    )

facility_status = {
    'Animation Room': 'Not Clocked In',
    'Coloring Area': 'Not Clocked In',
    'Lawn Activity Zone': 'Not Clocked In',
    'Cultural Activity Room': 'Not Clocked In',
}

@app.route('/facilities_usage', methods=['POST', 'GET'])
def facilities_usage():
    try:
        message = None

        if request.method == 'POST':
            facility_name = request.form['facility_name']
            action = request.form['action']

            current_status = facility_status[facility_name]

            if action == 'Clock In' and current_status == 'Not Clocked In':
                facility_status[facility_name] = 'Clocked In'
                message = f"Clocked in to {facility_name}."
            elif action == 'Clock Out' and current_status == 'Clocked In':
                facility_status[facility_name] = 'Not Clocked In'
                message = f"Clocked out from {facility_name}."
            else:
                message = f"Cannot perform {action} for {facility_name}. Facility is already in use."

        return render_template('facilities_usage.html', user_role=userrole, message=message)
    except Exception as e:
        return f"An error occurred: {str(e)}"

def open_browser(url):
    # Open the URL in the default web browser
    webbrowser.open(url)

if __name__ == '__main__':
    app.secret_key = 'super_secret_key'  # Change this secret key in a production environment
    # Specify the URL you want to open
    target_url = "http://127.0.0.1:5000"
    # Call the function to open the browser with the specified URL
    open_browser(target_url)
    app.run(debug=False)


